import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

file_path = "NS.xlsx"
sheet_name = "Hoja 1"

print("Reading Excel file...")

# --- 1. Try to detect the real header row automatically ---
df = None
for header_row in range(5):  # Try first 5 rows
    temp = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    cols = [str(c).strip() for c in temp.columns]
    if any("notes" in c.lower() for c in cols) and any("diag" in c.lower() for c in cols):
        df = temp
        break

if df is None:
    raise ValueError("Could not detect header row. Check Excel formatting (headers may start below row 5).")

# --- 2. Clean up column names ---
df.columns = [
    " ".join(map(str, c)).strip() if isinstance(c, (list, tuple)) else str(c).strip()
    for c in df.columns
]
df.columns = [c.replace("\xa0", " ").strip() for c in df.columns]  # remove non-breaking spaces

# --- 3. Helper to find the correct column names dynamically ---
def find_col(possible_names):
    for p in possible_names:
        for c in df.columns:
            if p.lower() == c.lower() or p.lower() in c.lower():
                return c
    return None

col_notes = find_col(["Notes"])
col_req = find_col(["Req Doc Diag Name"])
col_title = find_col(["Diagnostic Title"])
col_app = find_col(["Diagnostic Applicability"])

found = [col_notes, col_req, col_title, col_app]
if None in found:
    print("Missing expected columns. Found columns:")
    for c in df.columns:
        print(" -", c)
    raise ValueError("Could not find all required columns automatically.")

# --- 4. Subset and standardize column names ---
df = df[[col_notes, col_req, col_title, col_app]].rename(columns={
    col_notes: "Notes",
    col_req: "Req Doc Diag Name",
    col_title: "Diagnostic Title",
    col_app: "Diagnostic Applicability"
})

# --- 5. Focus only on INF/IUPR diagnostics ---
df = df[df.apply(lambda r: r.astype(str).str.contains("INF|IUPR", case=False).any(), axis=1)]

# --- 6. Extract relationships from the Notes text ---
def extract_relationships(text):
    if not isinstance(text, str):
        return None
    text = text.replace("\n", " ").strip()
    merge = re.findall(r"(?:combine|merge|join)\s+((?:INF|IUPR)\d+).*?(?:with|and)\s+((?:INF|IUPR)\d+)", text, re.I)
    keep = re.findall(r"(?:keep|retain)\s+((?:INF|IUPR)\d+)", text, re.I)
    delete = re.findall(r"(?:delete|remove)\s+((?:INF|IUPR)\d+)", text, re.I)
    return {"merge": merge, "keep": keep, "delete": delete}

df["Relationships"] = df["Notes"].apply(extract_relationships)

# --- 7. Filter only rows with actual merge/delete/keep relationships ---
def has_relationship(x):
    return isinstance(x, dict) and any(x.get(k) for k in ["merge", "keep", "delete"])

rel_df = df[df["Relationships"].apply(has_relationship)]

# --- 8. Build the summary records ---
records = []
for _, r in rel_df.iterrows():
    rel = r["Relationships"]
    name = str(r["Req Doc Diag Name"]).strip()
    title = str(r["Diagnostic Title"]).strip()
    app = str(r["Diagnostic Applicability"]).strip()

    if rel.get("merge"):
        for a, b in rel["merge"]:
            records.append({
                "Merge Group": f"{a} + {b}",
                "Req Doc Diag Name(s)": f"{a}, {b}",
                "Remaining": a,
                "Eliminated": b,
                "Diagnostic Title": title,
                "Diagnostic Applicability": app
            })
    if rel.get("delete"):
        for d in rel["delete"]:
            records.append({
                "Merge Group": d,
                "Req Doc Diag Name(s)": name,
                "Remaining": "",
                "Eliminated": d,
                "Diagnostic Title": title,
                "Diagnostic Applicability": app
            })
    if rel.get("keep"):
        for k in rel["keep"]:
            records.append({
                "Merge Group": k,
                "Req Doc Diag Name(s)": name,
                "Remaining": k,
                "Eliminated": "",
                "Diagnostic Title": title,
                "Diagnostic Applicability": app
            })

summary = pd.DataFrame(records)
if summary.empty:
    print("No INF/IUPR relationships found.")
    exit()

# --- 9. Collapse duplicates safely ---
def flatten_unique(values):
    flat = []
    for v in values:
        if isinstance(v, (list, tuple)):
            flat.extend(v)
        elif isinstance(v, str):
            flat.extend([s.strip() for s in v.split(",") if s.strip()])
    return sorted(set(flat) - {""})

summary = summary.groupby("Merge Group", as_index=False).agg({
    "Req Doc Diag Name(s)": lambda x: ", ".join(flatten_unique(x)),
    "Remaining": lambda x: ", ".join(flatten_unique(x)),
    "Eliminated": lambda x: ", ".join(flatten_unique(x)),
    "Diagnostic Title": "first",
    "Diagnostic Applicability": "first"
})

# --- 10. Save to Excel ---
out_file = "INF_IUPR_summary.xlsx"
summary.to_excel(out_file, index=False)

# --- 11. Beautify the Excel output ---
wb = load_workbook(out_file)
ws = wb.active

# Header styling
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Row styling and auto-sizing
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical="top")

for col in ws.columns:
    max_length = max(len(str(c.value)) if c.value else 0 for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)

wb.save(out_file)

# --- 12. Display result summary ---
print(f"\n✅ Summary generated successfully: {out_file}\n")
print("=== Quick Copy-Paste Table (tab-separated) ===\n")
print(summary.to_csv(sep="\t", index=False))
